from openpyxl import load_workbook


def read_excel_to_dict(path: str) -> dict[str, list[dict[str, str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {}

    for ws in wb.worksheets:
        sheet_data = []

        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            result[ws.title] = []
            continue

        header_row = rows[0]

        # 过滤掉空列名
        valid_indices = [
            idx for idx, col in enumerate(header_row)
            if col is not None and str(col).strip() != ""
        ]

        headers = [str(header_row[i]).strip() for i in valid_indices]

        # 读取数据行
        for row in rows[1:]:
            row_dict = {}
            for col_idx, header in zip(valid_indices, headers):
                value = row[col_idx] if col_idx < len(row) else None
                row_dict[header] = "" if value is None else str(value)
            sheet_data.append(row_dict)

        result[ws.title] = sheet_data

    wb.close()
    return result


def update_cells(path: str, sheet_name: str, updates: dict[str, str]):
    """
    updates 示例:
    {
        "A2": "新内容",
        "C5": "修改值",
    }
    """

    wb = load_workbook(path, keep_vba=True)
    ws = wb[sheet_name]

    for cell_addr, new_value in updates.items():
        ws[cell_addr].value = new_value  # 只改 value

    wb.save(path)
    wb.close()


def update_by_key(path, sheet_name, key_col, key_value, target_col, new_value):
    """
    找到 key_col 的值为 key_value 的行，将该行 target_col 的值改为 new_value
    key_col 和 target_col 都是列名
    """
    wb = load_workbook(path, keep_vba=True)
    ws = wb[sheet_name]

    # 找列号
    header = [cell.value for cell in ws[1]]

    key_idx = header.index(key_col) + 1
    target_idx = header.index(target_col) + 1

    for row in ws.iter_rows(min_row=2):
        # 找到对应的行
        if row[key_idx - 1].value == key_value:
            ws.cell(row=row[0].row, column=target_idx).value = new_value
            break

    wb.save(path)
    wb.close()
